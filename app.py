import os
import random
from flask import Flask, request, render_template
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/quiz', methods=['POST'])
def quiz():
    file = request.files['file']
    quiz_type = request.form.get('quizType')
    order = request.form.get('order', 'sequential')
    start_number = int(request.form.get('start-number', 1))

    # 读取 Excel 文件
    workbook = load_workbook(file)
    sheet = workbook.active
    questions_raw = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # 根据用户选择的题型筛选题目
    if quiz_type != 'all':
        if quiz_type == '判断题':
            quiz_type = '判断'
        elif quiz_type == '单选题':
            quiz_type = '单选'
        elif quiz_type == '多选题':
            quiz_type = '多选'

        questions_raw = [row for row in questions_raw if row[1] == quiz_type]

    if order == 'random':
        random.shuffle(questions_raw)
    else:
        questions_raw = questions_raw[start_number - 1:]

    # 截取前100道题
    questions_raw = questions_raw[:100]
    
    questions = []
    for question_data in questions_raw:
        question = {
            "题目类型": question_data[1],
            "序号": question_data[2],
            "题目": question_data[3],
            "选项A": question_data[4],
            "选项B": question_data[5],
            "选项C": question_data[6],
            "选项D": question_data[7],
            "正确答案": [c for c in question_data[8] if c != ' '] if question_data[8] is not None else [],  # 检查 None 值并跳过
        }
        questions.append(question)

    return render_template('quiz.html', questions=questions)

if __name__ == '__main__':
    app.run(debug=True)